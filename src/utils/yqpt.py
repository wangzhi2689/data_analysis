import datetime
import os
import shutil

import openpyxl
import pandas as pd
from collections import defaultdict

def read_info_file(info_file):
    row_data = defaultdict(list)
    data_xls = pd.io.excel.ExcelFile(info_file)
    try:
        for sheet_name in data_xls.sheet_names:
            if sheet_name != "info":
                continue
            output = pd.read_excel(info_file, engine='openpyxl', sheet_name=sheet_name)
            print(sheet_name + " read end, total of " + str(output.shape[0]) + " 行 " + str(output.shape[1]) + " 列.")
            workbook = openpyxl.load_workbook(info_file)
            sheet_obj = workbook[sheet_name]
            for c_num in range(2, output.shape[1]+1):
                for r_num in range(1, output.shape[0]):
                    tmpKey = str(sheet_obj.cell(5, c_num).value) + "-" + str(sheet_obj.cell(6, c_num).value)
                    if r_num == 12 or r_num == 16 or r_num == 17:
                        tmp = str(sheet_obj.cell(r_num, c_num).value).split()[0]
                        row_data[tmpKey].append(tmp)
                    else:
                        row_data[tmpKey].append(str(sheet_obj.cell(r_num, c_num).value))
        for key in row_data.keys():
            # print(key, row_data[key])
            pass
            for v in row_data[key]:
                # print(key, v)
                pass
    except Exception as e:
        print(e)
    finally:
        print(str(info_file) + "read_info_file end!")
    return row_data

#xx月xx日xx街道xx社区环境采样消杀
def update_excel_file(readFile, row_data):

    data_xls = pd.io.excel.ExcelFile(readFile)
    cur_month = str(datetime.datetime.now().month)
    cur_day = str(datetime.datetime.now().day)
    res_file = cur_month + "月" + cur_day +"日太华路街道纱厂东街社区环境采样消杀.xlsx"
    shutil.copyfile(readFile, res_file)
    try:
        for sheet_name in data_xls.sheet_names:
            output = pd.read_excel(readFile, engine='openpyxl', sheet_name=sheet_name)
            print(sheet_name + " read end, 共 " + str(output.shape[0]) + " 行 " + str(output.shape[1]) + " 列.")
            workbook = openpyxl.load_workbook(readFile)
            sheet_obj = workbook[sheet_name]
            for r_num in range(3, len(row_data) + 3):
                for key,values in row_data.items():
                    c_num_tmp = 0
                    for value in values:
                        if not value or str.isspace(value) or value =="None":
                            continue
                        c_num_tmp = c_num_tmp + 1
                        sheet_obj.cell(r_num, c_num_tmp).value = value

            workbook.save(res_file)
    except Exception as e:
        print(e)
    finally:
        print(str(readFile) + " read end!")
    return row_data

#体温表
def update_excel_file_2(readFile, row_data):

    data_xls = pd.io.excel.ExcelFile(readFile)
    cur_month = str(datetime.datetime.now().month)
    cur_day = str(datetime.datetime.now().day)
    res_file = cur_month + "月" + cur_day +"体温表消杀门扣表.xlsx"
    shutil.copyfile(readFile, res_file)
    try:
        for sheet_name in data_xls.sheet_names:
            if "体温表" != sheet_name:
                continue
            output = pd.read_excel(readFile, engine='openpyxl', sheet_name=sheet_name)
            print(sheet_name + " read end, 共 " + str(output.shape[0]) + " 行 " + str(output.shape[1]) + " 列.")
            workbook = openpyxl.load_workbook(readFile)
            sheet_obj = workbook[sheet_name]
            for r_num in range(1, output.shape[0] + 1):
                if(r_num == 1 or r_num == 2 or r_num == 4 or r_num == 6):
                    continue

                for key1,values in row_data.items():
                    print("333333333")
                    print(key1, values)
                    for value in values:
                        if not value or str.isspace(value) or value =="None":
                            continue

                        sheet_obj.cell(3, 1).value = row_data[key1][11]
                        sheet_obj.cell(3, 2).value = row_data[key1][4]
                        sheet_obj.cell(3, 3).value = row_data[key1][12]
                        sheet_obj.cell(3, 4).value = row_data[key1][13]
                        sheet_obj.cell(3, 5).value = row_data[key1][14]
                        sheet_obj.cell(3, 6).value = row_data[key1][15]
                        sheet_obj.cell(3, 7).value = row_data[key1][16]

                        sheet_obj.cell(5, 1).value = row_data[key1][17]
                        sheet_obj.cell(5, 3).value = row_data[key1][18]
                        sheet_obj.cell(5, 5).value = row_data[key1][6]

                        start_date = datetime.datetime.strptime(row_data[key1][15], "%Y-%m-%d")
                        for index in range(0, 100):
                            end_date = start_date + datetime.timedelta(days=index)
                            print(end_date.strftime("%Y-%m-%d") + " ---> " + row_data[key1][15])
                            sheet_obj.cell(index + 7, 2).value = end_date.strftime("%Y-%m-%d")
                            if end_date.strftime("%Y-%m-%d") == row_data[key1][16]:
                                break

            workbook.save(res_file)
    except Exception as e:
        print(e)
    finally:
        print(str(readFile) + " read end!")
    return row_data

def get_age_from_card_num(num):
    pass
def skip_row_data(sheet_obj, r_num, c_num_total):
    invild_str = ['序号', 'None', '日期', '身份证号']
    for c_num in range(1, c_num_total):
        cell = sheet_obj.cell(row=r_num, column=c_num)
        if type(cell).__name__ == 'MergedCell':
            print("Oh no, the cell is merged!")
            return True
        if str(sheet_obj.cell(r_num, c_num).value) in invild_str and c_num < 2:
            return True
    return False

def write_excel_file(data, writeFile):
    isSuccess = True
    try:
        writer = pd.ExcelWriter(writeFile)
        for sheet_name in data:
            df = pd.DataFrame(data[sheet_name])
            df.to_excel(excel_writer=writer, sheet_name=sheet_name)
    except Exception as e:
        print(e)
        isSuccess = False
    finally:
        writer.save()
        writer.close()
        print(str(writeFile) + " write end!" + str(isSuccess))
    return isSuccess

if __name__ == '__main__':

    import argparse
    parser = argparse.ArgumentParser(description="excel_utils.py usage help document")
    # 添加不带默认值的可解析参数
    parser.add_argument("-r", "--readFile", help="read excel file")  # 注意： -h、--help为内置参数，不可用
    ARGS = parser.parse_args()  # 获取命令行参数
    print('ARGS:', ARGS)

    if(not ARGS.readFile):
        row_data = defaultdict(list)
        row_data = read_info_file("信息登记.xlsx")
        print(row_data)
        data1 = update_excel_file("xx月xx日xx街道xx社区环境采样消杀.xlsx", row_data)
        data2 = update_excel_file_2("体温表消杀门扣表.xlsx", row_data) # 体温表

    #dic = defaultdict(list)
    #write_excel_file(data, ARGS.writeFile)